from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from pathlib import Path

# Caminhos base
BASE_DIR = Path(__file__).resolve().parent.parent
INPUT_FILE = BASE_DIR / "data" / "input" / "dados.txt"
OUTPUT_FILE = BASE_DIR / "data" / "output" / "PetLove_OUTUBRO.xlsx"

# Cria workbook e aba
wb = Workbook()
ws = wb.active
ws.title = "PetLove_OUTUBRO"

# Cabeçalhos da planilha
headers = [
    "Data", "Paciente", "Tutor", "Procedimento", "Repasse PetLove", "Co. part.",
    "TOTAL", "PAGO", "Custo", "Resultado", "35% do Resultado"
]
ws.append(headers)

# Estilo do cabeçalho
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Estilo colunas amarelas
highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Leitura do arquivo TXT
if not INPUT_FILE.exists():
    print(f"❌ Arquivo {INPUT_FILE} não encontrado!")
else:
    with open(INPUT_FILE, "r", encoding="utf-8") as f:
        linhas = [linha.strip() for linha in f if linha.strip()]

    if not linhas:
        print("⚠️ O arquivo TXT está vazio!")
    else:
        for linha in linhas:
            valores = linha.split(",")

            # Proteção: se tiver menos colunas
            while len(valores) < 6:
                valores.append("")

            # Converte valores numéricos
            try:
                repasse = float(valores[4])
                copart = float(valores[5])
            except ValueError:
                repasse = 0
                copart = 0

            total = repasse + copart
            resultado = total  # ou qualquer outra lógica
            resultado_35 = resultado * 0.35

            ws.append(valores + [
                f"R${total:.2f}", "-", "R$", f"R${resultado:.2f}", f"R${resultado_35:.2f}"
            ])

        # Aplica o amarelo nas colunas de destaque
        for row in ws.iter_rows(min_row=2, min_col=7, max_col=11):
            for cell in row:
                cell.fill = highlight_fill
                cell.alignment = Alignment(horizontal="center")

        # Ajusta largura das colunas
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18

        # Salva planilha
        wb.save(OUTPUT_FILE)
        print(f"✅ Planilha gerada com sucesso em: {OUTPUT_FILE}")
